"""
format_detector.py - Auto-detect energy meter file format from column headers.

Returns a cfg dict compatible with clean_data.process_format().
Returns None when the format cannot be identified.
"""

import csv
import io
import re
from pathlib import Path
from typing import Optional

import pandas as pd


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _file_type(filename: str) -> str:
    return {".xlsx": "xlsx", ".xls": "xlsx", ".xlsb": "xlsb", ".tsv": "tsv"}.get(
        Path(filename).suffix.lower(), "csv"
    )


def _read_sample(file_bytes: bytes, filename: str, nrows: int = 6, header_row: int = 0):
    """Return (df, raw_rows).  df may be None on read error.  raw_rows is CSV only."""
    suffix = Path(filename).suffix.lower()
    raw_rows = None
    try:
        if suffix in (".xlsx",):
            df = pd.read_excel(
                io.BytesIO(file_bytes), dtype=object,
                nrows=nrows, engine="openpyxl", header=header_row,
            )
            return df, raw_rows
        if suffix in (".xls",):
            df = pd.read_excel(
                io.BytesIO(file_bytes), dtype=object,
                nrows=nrows, engine="xlrd", header=header_row,
            )
            return df, raw_rows
        if suffix in (".xlsb", ".tsv"):
            df = pd.read_csv(
                io.BytesIO(file_bytes), sep="\t", dtype=object,
                nrows=nrows, comment="#", encoding="utf-8", errors="replace",
                header=header_row,
            )
            return df, raw_rows
        # CSV / unknown: also capture raw rows for NEM12 marker check
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        raw_rows = [row for i, row in enumerate(reader) if i < nrows + 2]
        df = pd.read_csv(
            io.StringIO(text), dtype=object, nrows=nrows, header=header_row,
        )
        return df, raw_rows
    except Exception:
        return None, raw_rows


def _norm_map(df: pd.DataFrame) -> dict:
    """Return {normalised_lower_stripped: original_col_name}."""
    result = {}
    for c in df.columns:
        key = str(c).lower().strip().lstrip("﻿")
        result[key] = str(c)
    return result


def _get(nm: dict, *candidates: str) -> Optional[str]:
    """Return the original column name for the first matching candidate (case-insensitive)."""
    for c in candidates:
        found = nm.get(c.lower().strip())
        if found is not None:
            return found
    return None


def _contains(nm: dict, fragment: str) -> Optional[str]:
    """Return the first original column name whose normalised form contains fragment."""
    frag = fragment.lower()
    for key, orig in nm.items():
        if frag in key:
            return orig
    return None


def _extract_nmi(filename: str) -> str:
    stem = Path(filename).stem
    m = re.search(r"[Nn]\d{9}|\d{10}", stem)
    return m.group(0)[:10].upper() if m else ""


def _is_ampm(df: pd.DataFrame, col: str) -> bool:
    """True if the datetime column values look like AM/PM (12-hour clock)."""
    if col not in df.columns:
        return False
    for val in df[col].dropna().head(5):
        if re.search(r"\b(am|pm)\b", str(val), re.IGNORECASE):
            return True
    return False


def _is_nem12_xlsx(file_bytes: bytes) -> bool:
    """Peek at an xlsx to see if its first two cells are '100' and 'NEM12'."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True, max_row=2))
        wb.close()
        if rows and rows[0]:
            row0 = [str(v).strip() for v in rows[0] if v is not None]
            if len(row0) >= 2 and row0[0] == "100" and row0[1] == "NEM12":
                return True
        return False
    except Exception:
        return False


def _is_excel_serial(df: pd.DataFrame, col: str) -> bool:
    """True if a column holds Excel date serial numbers (floats ~40000-60000)."""
    if col not in df.columns:
        return False
    for val in df[col].dropna().head(5):
        try:
            f = float(str(val))
            if 40000 < f < 60000:
                return True
        except (ValueError, TypeError):
            pass
    return False


def _detect_interval(df: pd.DataFrame, col: str) -> int:
    """Estimate interval in minutes from values in a datetime column."""
    if col not in df.columns:
        return 30
    import pandas as _pd
    for val in df[col].dropna().head(8):
        if isinstance(val, _pd.Timestamp):
            m = val.minute
        elif hasattr(val, "minute"):
            m = val.minute
        else:
            mch = re.search(r":(\d{2})(?::\d{2})?", str(val).strip())
            if not mch:
                continue
            m = int(mch.group(1))
        if m % 30 != 0 and m % 5 == 0:
            return 5
    return 30


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def detect_format(file_bytes: bytes, filename: str) -> Optional[dict]:
    """
    Detect energy meter format from file bytes + filename.
    Returns a cfg dict for clean_data.process_format(), or None if undetected.
    """
    ft = _file_type(filename)

    # ── NEM12 xlsx ──────────────────────────────────────────────────────────
    if ft == "xlsx" and _is_nem12_xlsx(file_bytes):
        return {
            "file_type": "xlsx", "layout": "nem12", "interval_minutes": 30,
            "datetime_col": "_nem12_", "ce_col": "_nem12_",
        }

    df, raw_rows = _read_sample(file_bytes, filename)

    # ── NEM12 CSV (standard: starts with 100,NEM12) ──────────────────────────
    if raw_rows:
        for row in raw_rows[:3]:
            if (len(row) >= 2
                    and str(row[0]).strip() == "100"
                    and str(row[1]).strip() == "NEM12"):
                return {
                    "file_type": ft, "layout": "nem12", "interval_minutes": 30,
                    "datetime_col": "_nem12_", "ce_col": "_nem12_",
                }

    # ── NEM12 CSV without 100 header (starts at 200 record) ─────────────────
    if raw_rows and raw_rows[0] and str(raw_rows[0][0]).strip() == "200":
        r = raw_rows[0]
        if (len(r) > 2
                and re.match(r"^[A-Z0-9]{8,12}$", str(r[1]).strip(), re.I)
                and re.search(r"[BE]1", str(r[2]).strip())):
            return {
                "file_type": ft, "layout": "nem12", "interval_minutes": 30,
                "datetime_col": "_nem12_", "ce_col": "_nem12_",
            }

    # ── Wide daily CSV: NMI/metadata rows + Date/Time row + HH:MM interval cols
    if raw_rows and len(raw_rows) > 3:
        r0, r3 = raw_rows[0], raw_rows[3]
        if (r0 and str(r0[0]).strip() == "NMI"
                and r3 and str(r3[0]).strip() == "Date/Time"
                and len(r3) > 2
                and re.match(r"^\d{1,2}:\d{2}$", str(r3[1]).strip())):
            nmi = str(r0[1]).strip()[:10] if len(r0) > 1 else ""
            return {
                "file_type": ft,
                "layout": "wide_daily",
                "const_nmi": nmi,
                "datetime_col": "Date/Time",
                "datetime_kind": "string_start",
                "ce_col": "_wide_daily_",
                "header_row": 3,
                "interval_minutes": 30,
            }

    # ── NMI metadata row 0 + offset header with endtime + kWh ──────────────
    if raw_rows and len(raw_rows) >= 2:
        r0 = raw_rows[0]
        if r0 and str(r0[0]).strip().upper() == "NMI" and len(r0) > 1:
            nmi_val = str(r0[1]).strip()[:10]
            df1, _ = _read_sample(file_bytes, filename, nrows=4, header_row=1)
            if df1 is not None and not df1.empty:
                nm1     = _norm_map(df1)
                et_col  = _get(nm1, "endtime")
                kwh_col = _get(nm1, "kwh")
                if et_col and kwh_col:
                    return {
                        "file_type": ft,
                        "header_row":    1,
                        "const_nmi":     nmi_val,
                        "datetime_col":  et_col,
                        "datetime_kind": "string_end",
                        "ce_col":        kwh_col,
                        "interval_minutes": 30,
                    }

    # ── Export Period CSV (offset header: NMI in metadata, header at row N) ──
    if raw_rows and raw_rows[0] and "export period" in str(raw_rows[0][0]).lower():
        header_idx = None
        nmi_val = ""
        for i, row in enumerate(raw_rows):
            if not row:
                continue
            cell0 = str(row[0]).strip()
            if not nmi_val:
                m = re.search(r"NMI[:\s]+([\w\-]+)", cell0, re.IGNORECASE)
                if m:
                    nmi_val = m.group(1)[:10]
            if cell0.lower() == "date time":
                header_idx = i
                break
        if header_idx is not None:
            return {
                "file_type": ft,
                "header_row": header_idx,
                "const_nmi": nmi_val or _extract_nmi(filename),
                "datetime_col": "Date Time",
                "datetime_kind": "string_end",
                "ce_col": "kWh",
                "interval_minutes": 30,
            }

    if df is None or df.empty:
        return None

    nm = _norm_map(df)
    cols = set(nm.keys())

    # ── GMT Time Adj xlsx (multi-row metadata, header at row with "Date" label) ─
    if ft == "xlsx" and _contains(nm, "gmt time adj"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            header_idx = None
            nmi_val = ""
            for i, row in enumerate(ws.iter_rows(values_only=True, max_row=25)):
                if not row:
                    continue
                c0 = str(row[0]).strip() if row[0] is not None else ""
                c1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if c0 == "NMI:":
                    nmi_val = c1[:10]
                if c0.lower() == "date":
                    header_idx = i
                    break
            wb.close()
            if header_idx is not None:
                df1, _ = _read_sample(file_bytes, filename, nrows=3, header_row=header_idx)
                if df1 is not None and not df1.empty:
                    nm1 = _norm_map(df1)
                    ce_col = _contains(nm1, "consumption")
                    return {
                        "file_type": ft,
                        "header_row": header_idx,
                        "const_nmi": nmi_val or _extract_nmi(filename),
                        "datetime_col": "Date",
                        "datetime_kind": "string_end",
                        "ce_col": ce_col,
                        "interval_minutes": 30,
                    }
        except Exception:
            pass

    # ── All-Unnamed headers: real header is 2 rows down (e.g. 51463 xlsx) ────
    all_unnamed = df.columns.size > 0 and all("unnamed" in str(c).lower() for c in df.columns)
    if all_unnamed and len(df) >= 2:
        row1_vals = {str(v).strip().lower() for v in df.iloc[1].values
                     if str(v).strip().lower() not in ("nan", "none", "")}
        if "nmi" in row1_vals and "kwh" in row1_vals:
            df2, _ = _read_sample(file_bytes, filename, nrows=4, header_row=2)
            if df2 is not None and not df2.empty:
                nm2 = _norm_map(df2)
                nmi_c  = _get(nm2, "nmi")
                dt_c   = _get(nm2, "interval")
                kwh_c  = _get(nm2, "kwh")
                if nmi_c and dt_c and kwh_c:
                    return {
                        "file_type": ft,
                        "header_row": 2,
                        "nmi_col":       nmi_c,
                        "datetime_col":  dt_c,
                        "datetime_kind": "string_end",
                        "ce_col":        kwh_c,
                        "interval_minutes": 30,
                    }

    # ── 1. Gentrack standard (nmi + endtime + kwh) ───────────────────────────
    if {"nmi", "endtime", "kwh"}.issubset(cols):
        cfg = {
            "file_type": ft,
            "nmi_col":      _get(nm, "nmi"),
            "datetime_col": _get(nm, "endtime"),
            "datetime_kind": "string_end",
            "ce_col":       _get(nm, "kwh"),
            "interval_minutes": 30,
        }
        soe = _get(nm, "generated_kwh")
        if soe:
            cfg["soe_col"] = soe
        return cfg

    # ── 2. Gentrack no-NMI (endtime + kwh, NMI from meter_serial or filename) ─
    endtime_col = _get(nm, "endtime", "end time")
    if endtime_col and _get(nm, "kwh") and "nmi" not in cols:
        cfg = {
            "file_type": ft,
            "datetime_col": endtime_col,
            "datetime_kind": "string_end",
            "ce_col":       _get(nm, "kwh"),
            "interval_minutes": 30,
        }
        meter_nmi = _get(nm, "meternmi", "meter nmi")
        if meter_nmi:
            cfg["nmi_col"] = meter_nmi
        else:
            cfg["const_nmi"] = _extract_nmi(filename)
        soe = _get(nm, "generated_kwh", "generated kwh")
        if soe:
            cfg["soe_col"] = soe
        qual = _get(nm, "quality", "quality code")
        if qual:
            cfg["quality_col"] = qual
        return cfg

    # ── MarketId + Time + Consumption kWh (e.g. 7001351111.csv) ─────────────
    mkt_col = _get(nm, "marketid")
    cons_kwh = _contains(nm, "consumption kwh")
    if mkt_col and cons_kwh:
        gen_kwh = _contains(nm, "generation kwh")
        return {
            "file_type": ft,
            "nmi_col":      mkt_col,
            "datetime_col": _get(nm, "time") or _get(nm, "datetime"),
            "datetime_kind": "string_end",
            "ce_col":       cons_kwh,
            "soe_col":      gen_kwh,
            "interval_minutes": 30,
        }

    # ── ReadDateTime + SiteCode + Consumption(E) (DetailUsage) ──────────────
    rd_dt = _get(nm, "readdatetime")
    site_c = _get(nm, "sitecode")
    cons_e = _contains(nm, "consumption (e)") or _contains(nm, "consumption(e)")
    if rd_dt and site_c and cons_e:
        gen_b = _contains(nm, "generation (b)") or _contains(nm, "generation(b)")
        qual_c = _get(nm, "quality")
        return {
            "file_type": ft,
            "nmi_col":       site_c,
            "datetime_col":  rd_dt,
            "datetime_kind": "string_start",
            "ce_col":        cons_e,
            "soe_col":       gen_b,
            "quality_col":   qual_c,
            "interval_minutes": 30,
        }

    # ── 3. Origin / Retailer  ReadingDateTime + E (Usage kWh) ───────────────
    rdt_col  = _contains(nm, "readingdatetime")
    e_use_col = _contains(nm, "e (usage")
    b_gen_col = _contains(nm, "b (generation")
    if rdt_col and e_use_col:
        interval = 5 if _is_ampm(df, rdt_col) else 30
        cfg = {
            "file_type": ft,
            "datetime_col":  rdt_col,
            "datetime_kind": "string_start",
            "ce_col":        e_use_col,
            "interval_minutes": interval,
        }
        nmi_col = _get(nm, "nmi")
        if nmi_col:
            cfg["nmi_col"] = nmi_col
        else:
            cfg["const_nmi"] = _extract_nmi(filename)
        if b_gen_col:
            cfg["soe_col"] = b_gen_col
        return cfg

    # ── 4. Citipower  NMI + Read Datetime + E + B ───────────────────────────
    read_dt = _get(nm, "read datetime", "read_datetime")
    if read_dt and _get(nm, "nmi") and _get(nm, "e") and _get(nm, "b"):
        return {
            "file_type": ft,
            "nmi_col":      _get(nm, "nmi"),
            "datetime_col": read_dt,
            "datetime_kind": "string_start",
            "ce_col":       _get(nm, "e"),
            "soe_col":      _get(nm, "b"),
            "interval_minutes": 30,
        }

    # ── NMI + Read Datetime + kWh + B (AusNet style, string_start) ─────────
    rd2_col = _get(nm, "read datetime")
    if _get(nm, "nmi") and rd2_col and _get(nm, "kwh") and _get(nm, "b"):
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  rd2_col,
            "datetime_kind": "string_start",
            "ce_col":        _get(nm, "kwh"),
            "soe_col":       _get(nm, "b"),
            "interval_minutes": 30,
        }

    # ── NMI + Read Date/Time + kWh (EU dd.mm.yyyy format, 5-min) ───────────
    rdtime_col = _get(nm, "read date/time")
    if _get(nm, "nmi") and rdtime_col and _get(nm, "kwh"):
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  rdtime_col,
            "datetime_kind": "string_start",
            "ce_col":        _get(nm, "kwh"),
            "interval_minutes": _detect_interval(df, rdtime_col),
        }

    # ── NMI + REPORT_DATE + REPORT_TIME + KWH (MP Metering Provider, 5-min) ─
    if _get(nm, "nmi") and _get(nm, "report_date") and _get(nm, "report_time") and _get(nm, "kwh"):
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  _get(nm, "report_date"),
            "time_col":      _get(nm, "report_time"),
            "datetime_kind": "string_start",
            "ce_col":        _get(nm, "kwh"),
            "quality_col":   _get(nm, "quality"),
            "interval_minutes": 5,
        }

    # ── Identifier + Date + Start Time + Usage + Generation (Idameneo) ──────
    ident_col    = _get(nm, "identifier")
    id_type_col  = _get(nm, "identifiertype")
    start_time   = _get(nm, "start time")
    gen_col      = _get(nm, "generation")
    if ident_col and _get(nm, "date") and start_time and _get(nm, "usage") and gen_col:
        cfg = {
            "file_type": ft,
            "nmi_col":       ident_col,
            "datetime_col":  _get(nm, "date"),
            "time_col":      start_time,
            "datetime_kind": "string_start",
            "ce_col":        _get(nm, "usage"),
            "soe_col":       gen_col,
            "interval_minutes": 30,
        }
        if id_type_col:
            cfg["filter_col"]   = id_type_col
            cfg["filter_value"] = "NMI"
        return cfg

    # ── Commodity + Date + Time + NMI + Usage (e.g. 4310261111, 51481) ──────
    if _get(nm, "commodity") and _get(nm, "nmi") and _get(nm, "usage"):
        dt_c   = _get(nm, "date") or _get(nm, "datetime")
        tm_c   = _get(nm, "time")
        int_c  = _get(nm, "interval")
        iv_min = 30
        if int_c and not df.empty:
            m = re.match(r"(\d+)\s*min", str(df.iloc[0][int_c]).strip().lower())
            if m:
                iv_min = int(m.group(1))
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  dt_c,
            "time_col":      tm_c,
            "datetime_kind": "string_start",
            "ce_col":        _get(nm, "usage"),
            "interval_minutes": iv_min,
        }

    # ── Meter Code + Meter Register + Read Date + Meter Read (Shannon MDP) ───
    if (_get(nm, "meter code") and _get(nm, "meter register")
            and _get(nm, "read date") and _get(nm, "meter read")):
        rd_col = _get(nm, "read date")
        return {
            "file_type": ft,
            "layout":        "long",
            "const_nmi":     _extract_nmi(filename) or Path(filename).stem[:10],
            "datetime_col":  rd_col,
            "datetime_kind": "string_start",
            "type_col":      _get(nm, "meter register"),
            "ce_type_value":  "E1",
            "soe_type_value": "B1",
            "ce_col":        _get(nm, "meter read"),
            "quality_col":   _get(nm, "quality method"),
            "interval_minutes": _detect_interval(df, rd_col),
        }

    # ── meter_code + read_datetime + read_value (51460 MDP long format) ──────
    if _get(nm, "meter_code") and _get(nm, "read_datetime") and _get(nm, "read_value"):
        local_dst = _get(nm, "local_datetime_dst") or _get(nm, "local_datetime")
        dt_col    = local_dst or _get(nm, "read_datetime")
        return {
            "file_type": ft,
            "layout":        "long",
            "const_nmi":     _extract_nmi(filename) or Path(filename).stem[:10],
            "datetime_col":  dt_col,
            "datetime_kind": "string_end",
            "type_col":      _get(nm, "meter_code"),
            "ce_type_value":  "E1",
            "soe_type_value": "B1",
            "ce_col":        _get(nm, "read_value"),
            "interval_minutes": 30,
        }

    # ── NMI + DateTime + kWh only (simple 3-col ISO format, string_end) ────
    dt_only_col = _get(nm, "datetime")
    if _get(nm, "nmi") and dt_only_col and _get(nm, "kwh") and not _get(nm, "e") and not _get(nm, "b"):
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  dt_only_col,
            "datetime_kind": "string_end",
            "ce_col":        _get(nm, "kwh"),
            "interval_minutes": 30,
        }

    # ── NMI + DateTime + Usage (kWh) + Export (kWh) (51399 format) ──────────
    usage_kwh_col  = _contains(nm, "usage (kwh)")
    export_kwh_col = _contains(nm, "export (kwh)")
    if _get(nm, "nmi") and dt_only_col and usage_kwh_col:
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  dt_only_col,
            "datetime_kind": "string_end",
            "ce_col":        usage_kwh_col,
            "soe_col":       export_kwh_col,
            "interval_minutes": 30,
        }

    # ── 5. Citipower  NMI + Date (or Datetime) + E + B ──────────────────────
    if _get(nm, "nmi") and _get(nm, "e") and _get(nm, "b"):
        dt_col = _get(nm, "readingdatetime", "datetime", "date",
                       "local time", "starttime", "start time", "period")
        if dt_col:
            return {
                "file_type": ft,
                "nmi_col":      _get(nm, "nmi"),
                "datetime_col": dt_col,
                "datetime_kind": "string_start",
                "ce_col":       _get(nm, "e"),
                "soe_col":      _get(nm, "b"),
                "interval_minutes": 30,
            }

    # ── 6. Long: Nmi / Datetime / Meter Type / Active ────────────────────────
    active_col = _get(nm, "active")
    if _get(nm, "nmi") and _get(nm, "datetime") and _get(nm, "meter type") and active_col:
        return {
            "file_type": ft,
            "layout":        "long",
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  _get(nm, "datetime"),
            "datetime_kind": "string_start",
            "type_col":      _get(nm, "meter type"),
            "ce_type_value":  "E1",
            "soe_type_value": "B1",
            "ce_col":        active_col,
            "interval_minutes": 5,
        }

    # ── 7. Long: Nmi / Date / Time / Meter Type ──────────────────────────────
    if _get(nm, "nmi") and _get(nm, "date") and _get(nm, "time") and _get(nm, "meter type"):
        act = _get(nm, "active") or _contains(nm, "active")
        return {
            "file_type": ft,
            "layout":        "long",
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  _get(nm, "date"),
            "time_col":      _get(nm, "time"),
            "datetime_kind": "string_start",
            "type_col":      _get(nm, "meter type"),
            "ce_type_value":  "E1",
            "soe_type_value": "B1",
            "ce_col":        act,
            "interval_minutes": 5,
        }

    # ── NMI + Register + Date + Time + Energy(kWh) (Intellihub, type_startswith) ─
    energy_kwh = _contains(nm, "energy (kwh)")
    if (_get(nm, "nmi") and energy_kwh
            and _get(nm, "register") and _get(nm, "date") and _get(nm, "time")):
        int_sec_col = _contains(nm, "interval (seconds)")
        iv_min = 30
        if int_sec_col and not df.empty:
            try:
                iv_min = max(5, int(float(str(df.iloc[0][int_sec_col]).strip()) // 60))
            except Exception:
                pass
        return {
            "file_type": ft,
            "layout":          "long",
            "nmi_col":         _get(nm, "nmi"),
            "datetime_col":    _get(nm, "date"),
            "time_col":        _get(nm, "time"),
            "datetime_kind":   "string_start",
            "type_col":        _get(nm, "register"),
            "type_startswith": True,
            "ce_type_value":   "E1",
            "soe_type_value":  "B1",
            "ce_col":          energy_kwh,
            "interval_minutes": iv_min,
        }

    # ── Usage from Grid / Feed In (multi-section xlsx, real header at row 1) ──
    if _contains(nm, "usage from the grid") and _contains(nm, "feed in"):
        df1, _ = _read_sample(file_bytes, filename, nrows=8, header_row=1)
        if df1 is not None and not df1.empty:
            return {
                "file_type": ft,
                "header_row": 1,
                "data_skip_rows": 4,
                "const_nmi": _extract_nmi(filename) or Path(filename).stem[:10],
                "datetime_col": str(df1.columns[0]),
                "datetime_kind": "string_start",
                "ce_col": "Active Amt (kWh)",
                "soe_col": "Active Amt (kWh).1",
                "interval_minutes": 30,
            }

    # ── 8. Alba Thermal / Usage Data CSV  (Site + From Date + Direction + Unit + Value) ──
    from_dt = _get(nm, "from date", "from_date")
    if _get(nm, "site") and from_dt and _get(nm, "direction") and _get(nm, "unit") and _get(nm, "value"):
        cfg = {
            "file_type": ft,
            "layout":        "long",
            "filter_col":    _get(nm, "unit"),
            "filter_value":  "kWh",
            "nmi_col":       _get(nm, "site"),
            "datetime_col":  from_dt,
            "datetime_kind": "string_start",
            "type_col":      _get(nm, "direction"),
            "ce_type_value":  "I",
            "soe_type_value": "E",
            "ce_col":        _get(nm, "value"),
            "interval_minutes": 30,
        }
        from_tm = _get(nm, "from time", "from_time")
        if from_tm:
            cfg["time_col"] = from_tm
        return cfg

    # ── Direction + Interval Date/Time + Amount + UOM (long, IMPORT/EXPORT) ──
    int_dt_col = _contains(nm, "interval date")
    if int_dt_col and _get(nm, "direction") and _get(nm, "amount") and _get(nm, "uom"):
        return {
            "file_type": ft,
            "layout": "long",
            "filter_col": _get(nm, "uom"),
            "filter_value": "KWH",
            "const_nmi": _extract_nmi(filename) or Path(filename).stem[:10],
            "datetime_col": int_dt_col,
            "datetime_kind": "string_end",
            "type_col": _get(nm, "direction"),
            "ce_type_value": "IMPORT",
            "soe_type_value": "EXPORT",
            "ce_col": _get(nm, "amount"),
            "interval_minutes": 30,
        }

    # ── 9. EndDate + EndTime + Consumption ───────────────────────────────────
    if _get(nm, "enddate") and _get(nm, "endtime") and _get(nm, "consumption"):
        cfg = {
            "file_type": ft,
            "datetime_col":  _get(nm, "enddate"),
            "time_col":      _get(nm, "endtime"),
            "datetime_kind": "string_end",
            "ce_col":        _get(nm, "consumption"),
            "interval_minutes": 30,
        }
        nmi_col = _get(nm, "nmi")
        if nmi_col:
            cfg["nmi_col"] = nmi_col
        else:
            cfg["const_nmi"] = _extract_nmi(filename)
        uom_col = _get(nm, "consumptionuom") or _contains(nm, "uom")
        if uom_col:
            cfg["filter_col"]   = uom_col
            cfg["filter_value"] = "kWh"
        qual_col = _get(nm, "consumptionquality") or _contains(nm, "quality")
        if qual_col:
            cfg["quality_col"] = qual_col
        return cfg

    # ── NMI + StartDate + ProfileReadValue + RateTypeDescription (Jemena/DPE) ─
    profile_col  = _get(nm, "profilereadvalue")
    start_dt_col = _get(nm, "startdate", "start date", "startdatetime")
    rate_col     = _get(nm, "ratetypedescription")
    if _get(nm, "nmi") and start_dt_col and profile_col and rate_col:
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  start_dt_col,
            "datetime_kind": "string_start",
            "ce_col":        profile_col,
            "filter_col":    rate_col,
            "filter_value":  "Usage",
            "quality_col":   _get(nm, "qualityflag"),
            "interval_minutes": 30,
        }

    # ── 10. nmi + intervaldate/date + time/day_time + intervalvalue_k*h ──────
    iv_col = _contains(nm, "intervalvalue_k")
    date_c = _get(nm, "date") or _get(nm, "intervaldate")
    time_c = _get(nm, "time") or _get(nm, "day_time")
    if _get(nm, "nmi") and date_c and time_c and iv_col:
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  date_c,
            "time_col":      time_c,
            "datetime_kind": "string_end",
            "ce_col":        iv_col,
            "interval_minutes": 30,
        }

    # ── 11. Momentum / Service Point + Import kWh ────────────────────────────
    svc_col    = _get(nm, "service point", "service_point")
    import_col = _contains(nm, "import kwh") or _contains(nm, "import_kwh")
    if svc_col and import_col:
        dt_col     = (_get(nm, "period start", "date", "datetime", "read datetime")
                      or _contains(nm, "market date") or _contains(nm, "site date"))
        time_col   = _contains(nm, "market time") or _contains(nm, "site time")
        export_col = _contains(nm, "export kwh") or _contains(nm, "export_kwh")
        cfg = {
            "file_type": ft,
            "nmi_col":       svc_col,
            "datetime_col":  dt_col,
            "datetime_kind": "string_start",
            "ce_col":        import_col,
            "interval_minutes": 30,
        }
        if time_col:
            cfg["time_col"] = time_col
        if export_col:
            cfg["soe_col"] = export_col
        return cfg

    # ── 12. Read datetime + Export kWh / Import kWh ──────────────────────────
    rd_col  = _get(nm, "read datetime")
    imp_col = _contains(nm, "import kwh")
    exp_col = _contains(nm, "export kwh")
    if rd_col and imp_col:
        cfg = {
            "file_type": ft,
            "datetime_col":  rd_col,
            "datetime_kind": "string_end",
            "ce_col":        imp_col,
            "interval_minutes": 30,
        }
        nmi_col = _get(nm, "nmi") or _get(nm, "site") or svc_col
        if nmi_col:
            cfg["nmi_col"] = nmi_col
        else:
            cfg["const_nmi"] = _extract_nmi(filename)
        if exp_col:
            cfg["soe_col"] = exp_col
        return cfg

    # ── Datetime + Net Grid Import kWh (no NMI column, NMI from filename) ───
    net_import_col = _contains(nm, "net grid import")
    if _get(nm, "datetime") and net_import_col:
        return {
            "file_type": ft,
            "const_nmi": _extract_nmi(filename) or Path(filename).stem[:10],
            "datetime_col": _get(nm, "datetime"),
            "datetime_kind": "string_start",
            "ce_col": net_import_col,
            "interval_minutes": 30,
        }

    # ── MeterNMI + EST Time / Read Time + KWH+ (Mondo, 5-min) ───────────────
    meternmi_col = _get(nm, "meternmi")
    kwh_plus     = _get(nm, "kwh+")
    if meternmi_col and kwh_plus:
        dt_col   = _get(nm, "read time") or _get(nm, "est time") or _contains(nm, "time")
        kwh_min  = _get(nm, "kwh-")
        kind     = "excel_end" if _is_excel_serial(df, dt_col) else "string_end"
        return {
            "file_type": ft,
            "const_nmi":     _extract_nmi(filename) or Path(filename).stem[:10],
            "datetime_col":  dt_col,
            "datetime_kind": kind,
            "ce_col":        kwh_plus,
            "soe_col":       kwh_min,
            "interval_minutes": 5,
        }

    # ── Date + Channel + KWh (NEEE ColumnarCSV, long E1/B1) ─────────────────
    if _get(nm, "date") and _get(nm, "channel") and _get(nm, "kwh"):
        qual_c = _get(nm, "quality")
        return {
            "file_type": ft,
            "layout":        "long",
            "const_nmi":     _extract_nmi(filename) or Path(filename).stem[:10],
            "datetime_col":  _get(nm, "date"),
            "datetime_kind": "string_start",
            "type_col":      _get(nm, "channel"),
            "ce_type_value":  "E1",
            "soe_type_value": "B1",
            "ce_col":        _get(nm, "kwh"),
            "quality_col":   qual_c,
            "interval_minutes": 30,
        }

    # ── NMI + Local Time + E kWh at Meter (Kipand) ───────────────────────────
    e_kwh_at = _contains(nm, "e kwh")
    if _get(nm, "nmi") and _get(nm, "local time") and e_kwh_at:
        b_kwh_at = _contains(nm, "b kwh")
        return {
            "file_type": ft,
            "nmi_col":       _get(nm, "nmi"),
            "datetime_col":  _get(nm, "local time"),
            "datetime_kind": "string_start",
            "ce_col":        e_kwh_at,
            "soe_col":       b_kwh_at,
            "interval_minutes": 30,
        }

    # ── Date Time + kWh (no NMI col, NMI from filename) ──────────────────────
    date_time_col = _get(nm, "date time")
    if date_time_col and _get(nm, "kwh") and "nmi" not in cols:
        return {
            "file_type": ft,
            "const_nmi":     _extract_nmi(filename) or Path(filename).stem[:10],
            "datetime_col":  date_time_col,
            "datetime_kind": "string_end",
            "ce_col":        _get(nm, "kwh"),
            "interval_minutes": 30,
        }

    # ── Wide intervals: NMI + DATE + CONSUMPTION_TYPE + HH:MM-HH:MM cols ────
    if _get(nm, "nmi") and _get(nm, "date"):
        ct_col = _get(nm, "consumption_type") or _contains(nm, "consumption type")
        if ct_col:
            interval_cols = [c for c in df.columns
                             if re.match(r"\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}", str(c))]
            if len(interval_cols) >= 2:
                _CE  = {"e1", "e", "import", "consumption", "in", "delivered"}
                _SOE = {"b1", "b", "export", "generation", "out", "received", "feedin"}
                type_vals = (df[ct_col].dropna().astype(str).str.strip().unique().tolist()
                             if ct_col in df.columns else [])
                ce_val  = next((v for v in type_vals if v.lower() in _CE),  "E1")
                soe_val = next((v for v in type_vals if v.lower() in _SOE), None)
                return {
                    "file_type": ft,
                    "layout":        "wide_intervals",
                    "nmi_col":       _get(nm, "nmi"),
                    "datetime_col":  _get(nm, "date"),
                    "datetime_kind": "string_end",
                    "type_col":      ct_col,
                    "ce_type_value":  ce_val,
                    "soe_type_value": soe_val,
                    "filter_col":    _get(nm, "unit"),
                    "filter_value":  "KWH",
                    "quality_col":   _get(nm, "quality"),
                    "interval_minutes": 30,
                }

    # ── 13. NMI + E col (CE) + some datetime ────────────────────────────────
    if _get(nm, "nmi") and _get(nm, "e"):
        dt_col = _get(nm, "datetime", "date", "period", "timestamp")
        if dt_col:
            return {
                "file_type": ft,
                "nmi_col":      _get(nm, "nmi"),
                "datetime_col": dt_col,
                "datetime_kind": "string_start",
                "ce_col":       _get(nm, "e"),
                "soe_col":      _get(nm, "b"),
                "interval_minutes": 30,
            }

    # ── DateTime / Time Ending + NMI-prefixed energy columns (Yurika, 7105091111) ─
    dt_norm2 = next(
        (k for k in cols if k in ("datetime", "date time", "time ending")), None
    )
    if dt_norm2:
        nmi_prefix_cols = [c for c in df.columns if re.match(r"^\d{10}[^0-9]", str(c))]
        if nmi_prefix_cols:
            nmi = str(nmi_prefix_cols[0])[:10]
            import_c = next((c for c in nmi_prefix_cols if "import" in c.lower()), None)
            export_c = next((c for c in nmi_prefix_cols if "export" in c.lower()), None)
            kwh_c    = next((c for c in nmi_prefix_cols
                             if "kwh" in c.lower() and "demand" not in c.lower()), None)
            ce_col   = import_c or kwh_c or nmi_prefix_cols[0]
            iv_min   = _detect_interval(df, nm[dt_norm2])
            return {
                "file_type": ft,
                "const_nmi":     nmi,
                "datetime_col":  nm[dt_norm2],
                "datetime_kind": "string_end",
                "ce_col":        ce_col,
                "soe_col":       export_c,
                "interval_minutes": iv_min,
            }

    # ── 14. Wide NMIs: DateTime column + 3+ 10-digit NMI column headers ──────
    dt_norm = next(
        (k for k in cols if k in ("datetime", "date time", "timestamp", "date")), None
    )
    if dt_norm:
        nmi_header_cols = [
            k for k in cols
            if re.match(r"^\d{10}$", k) or re.match(r"^n\d{9}$", k.lower())
        ]
        if len(nmi_header_cols) >= 3:
            return {
                "file_type": ft,
                "layout":        "wide_nmis",
                "datetime_col":  nm[dt_norm],
                "datetime_kind": "string_start",
                "ce_col":        "_wide_nmis_",
                "interval_minutes": 30,
            }

    # ── Undetected ────────────────────────────────────────────────────────────
    return None
